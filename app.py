#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dashboard Server - حي السعادة 132
تطبيق ويب يقرأ بيانات من الإكسل
"""

from flask import Flask, render_template_string, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "dashboard_data.xlsx"

def get_dashboard_data():
    """قراءة بيانات Dashboard من الإكسل"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return get_default_data()
        
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['Dashboard']
        
        # استخراج البيانات
        total_sales = 0
        collected = 0
        collection_rate = 0
        completion = 0
        trust_total = 0
        total_units = 0
        sold_units = 0
        
        # قراءة الصفوف
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
            try:
                if row[1] and row[1].value:
                    label = str(row[1].value).strip()
                    
                    if "إجمالي المبيعات لكامل المشروع" in label and row[2]:
                        val = row[2].value
                        total_sales = float(val) if val else 0
                    
                    elif "الدفعات المحصلة" in label and row[2]:
                        val = row[2].value
                        collected = float(val) if val else 0
                    
                    elif "نسبة التحصيل من كامل المشروع" in label and row[2]:
                        val = row[2].value
                        collection_rate = float(val) * 100 if val else 0
                    
                    elif "نسبة الإنجاز" in label and row[2]:
                        val = row[2].value
                        completion = float(val) * 100 if val else 0
                    
                    elif "الإجمالي" in label and "الضمان" in label and row[4]:
                        val = row[4].value
                        trust_total = float(val) if val else 0
                    
                    elif "عدد الوحدات" in label and "المباعة" not in label and row[2]:
                        val = row[2].value
                        total_units = int(float(val)) if val else 0
                    
                    elif "عدد الوحدات المباعة" in label and row[2]:
                        val = row[2].value
                        sold_units = int(float(val)) if val else 0
            except:
                continue
        
        data = {
            'total_sales': total_sales,
            'collected': collected,
            'collection_rate': collection_rate,
            'completion': completion,
            'trust_total': trust_total,
            'charity': total_sales * 0.02,
            'total_units': total_units,
            'sold_units': sold_units,
            'available_units': total_units - sold_units,
            'remaining_sales': total_sales - collected,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return data
        
    except Exception as e:
        print(f"خطأ في قراءة الإكسل: {e}")
        return get_default_data()

def get_default_data():
    """بيانات افتراضية"""
    return {
        'total_sales': 14304084,
        'collected': 8415474.1,
        'collection_rate': 58.8,
        'completion': 56,
        'trust_total': 7928625.1,
        'charity': 286081.68,
        'total_units': 40,
        'sold_units': 25,
        'available_units': 15,
        'remaining_sales': 5888610,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

def format_number(num):
    """تنسيق الأرقام"""
    if num is None or num == 0:
        return "0"
    return f"{num:,.0f}"

@app.route('/api/data')
def api_data():
    """API لإرجاع البيانات JSON"""
    data = get_dashboard_data()
    return jsonify(data)

@app.route('/')
def dashboard():
    """الصفحة الرئيسية للداش بورد"""
    data = get_dashboard_data()
    
    html = f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>لوحة البيانات | قصر القصور</title>
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800;900&family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        
        :root {{
            --white: #FFFFFF;
            --cream: #F5F2EB;
            --gold: #B8915F;
            --gold-l: #D4AE7E;
            --gold-d: #8A6A3F;
            --slate: #2A2D33;
            --black: #0F1114;
            --text: #0F1114;
            --muted: #8B8F96;
            --line: #E8E5DD;
            --success: #4A7C5B;
            --success-bg: #E8F1EB;
            --danger: #B85959;
            --danger-bg: #F5E6E6;
        }}
        
        body {{
            font-family: "Cairo", sans-serif;
            background: linear-gradient(135deg, #F5F2EB 0%, #E8E2D2 100%);
            color: var(--text);
            direction: rtl;
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1420px;
            margin: 0 auto;
            background: var(--white);
            border-radius: 20px;
            overflow: hidden;
            box-shadow: 0 24px 80px rgba(15, 17, 20, 0.15);
            border: 1px solid var(--line);
        }}
        
        .topbar {{
            background: linear-gradient(135deg, var(--white) 0%, var(--cream) 100%);
            border-bottom: 1px solid var(--line);
            padding: 22px 32px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            position: relative;
        }}
        
        .topbar::before {{
            content: "";
            position: absolute;
            bottom: 0; right: 0; left: 0;
            height: 3px;
            background: linear-gradient(90deg, var(--gold) 0%, var(--gold-l) 50%, var(--gold) 100%);
        }}
        
        .brand {{
            display: flex;
            align-items: center;
            gap: 16px;
        }}
        
        .brand-logo {{
            width: 72px;
            height: 72px;
            border-radius: 14px;
            background: var(--white);
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 6px;
            box-shadow: 0 4px 14px rgba(15, 17, 20, 0.15);
            border: 1px solid var(--line);
            font-size: 40px;
        }}
        
        .brand-text h1 {{
            font-size: 20px;
            font-weight: 800;
            color: var(--black);
        }}
        
        .brand-text p {{
            font-size: 12px;
            color: var(--muted);
            margin-top: 4px;
        }}
        
        .topbar-badge {{
            background: linear-gradient(135deg, var(--gold), var(--gold-d));
            color: var(--white);
            font-size: 11px;
            font-weight: 700;
            padding: 7px 14px;
            border-radius: 24px;
            animation: pulse 2s infinite;
        }}
        
        @keyframes pulse {{
            0%, 100% {{ opacity: 1; }}
            50% {{ opacity: 0.7; }}
        }}
        
        .main {{
            padding: 28px;
            background: var(--cream);
        }}
        
        .hero {{
            background: linear-gradient(135deg, var(--black) 0%, var(--slate) 100%);
            border-radius: 18px;
            padding: 32px 36px;
            margin-bottom: 26px;
            color: var(--white);
        }}
        
        .hero h2 {{
            font-size: 32px;
            font-weight: 900;
            margin-bottom: 12px;
        }}
        
        .hero p {{
            font-size: 14px;
            opacity: 0.85;
            margin-bottom: 20px;
        }}
        
        .section-title {{
            font-size: 17px;
            font-weight: 800;
            color: var(--black);
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 12px;
        }}
        
        .section-title::before {{
            content: "";
            width: 5px;
            height: 22px;
            background: linear-gradient(180deg, var(--gold), var(--gold-d));
            border-radius: 3px;
        }}
        
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        
        .kpi-card {{
            background: var(--white);
            border: 1px solid var(--line);
            border-radius: 14px;
            padding: 20px;
            position: relative;
            overflow: hidden;
            transition: transform 0.25s;
        }}
        
        .kpi-card:hover {{
            transform: translateY(-3px);
            box-shadow: 0 12px 32px rgba(15, 17, 20, 0.1);
        }}
        
        .kpi-card::before {{
            content: "";
            position: absolute;
            top: 0; right: 0;
            width: 4px;
            height: 100%;
        }}
        
        .kpi-card.gold::before {{ background: linear-gradient(180deg, var(--gold-l), var(--gold-d)); }}
        .kpi-card.danger::before {{ background: linear-gradient(180deg, #D08585, var(--danger)); }}
        .kpi-card.success::before {{ background: linear-gradient(180deg, #5A9B6E, var(--success)); }}
        
        .kpi-icon {{
            font-size: 24px;
            margin-bottom: 12px;
        }}
        
        .kpi-label {{
            font-size: 12px;
            color: var(--muted);
            font-weight: 600;
            margin-bottom: 6px;
            text-transform: uppercase;
        }}
        
        .kpi-value {{
            font-size: 24px;
            font-weight: 900;
            color: var(--black);
            margin-bottom: 8px;
        }}
        
        .kpi-delta {{
            font-size: 12px;
            padding: 4px 8px;
            border-radius: 6px;
            font-weight: 600;
            display: inline-block;
        }}
        
        .kpi-delta.up {{ background: var(--success-bg); color: var(--success); }}
        .kpi-delta.neutral {{ background: var(--danger-bg); color: var(--danger); }}
        
        .card {{
            background: var(--white);
            border: 1px solid var(--line);
            border-radius: 14px;
            padding: 22px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        table tr {{
            border-bottom: 1px solid var(--line);
        }}
        
        table td {{
            padding: 12px;
            text-align: right;
        }}
        
        table td:first-child {{
            color: var(--muted);
            font-weight: 600;
        }}
        
        footer {{
            background: var(--cream);
            border-top: 1px solid var(--line);
            padding: 20px;
            text-align: center;
            font-size: 12px;
            color: var(--muted);
        }}
        
        @media (max-width: 768px) {{
            .kpi-grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="topbar">
            <div class="brand">
                <div class="brand-logo">🏢</div>
                <div class="brand-text">
                    <h1>قصر القصور</h1>
                    <p>نبني بجودة واحسان</p>
                </div>
            </div>
            <div class="topbar-badge">📊 Live</div>
        </div>
        
        <div class="main">
            <div class="hero">
                <h2>حي السعادة 132</h2>
                <p>مشروع سكني متطور في الرياض</p>
            </div>
            
            <div class="section-title">المؤشرات الرئيسية</div>
            <div class="kpi-grid">
                <div class="kpi-card danger">
                    <div class="kpi-icon">❤️</div>
                    <div class="kpi-label">الأعمال الخيرية</div>
                    <div class="kpi-value">{format_number(data['charity'])} ر.س</div>
                </div>
                
                <div class="kpi-card gold">
                    <div class="kpi-icon">📊</div>
                    <div class="kpi-label">إجمالي المبيعات</div>
                    <div class="kpi-value">{format_number(data['total_sales'])} ر.س</div>
                </div>
                
                <div class="kpi-card gold">
                    <div class="kpi-icon">💵</div>
                    <div class="kpi-label">الدفعات المحصلة</div>
                    <div class="kpi-value">{format_number(data['collected'])} ر.س</div>
                </div>
                
                <div class="kpi-card success">
                    <div class="kpi-icon">💰</div>
                    <div class="kpi-label">رصيد الضمان</div>
                    <div class="kpi-value">{format_number(data['trust_total'])} ر.س</div>
                </div>
                
                <div class="kpi-card gold">
                    <div class="kpi-icon">📈</div>
                    <div class="kpi-label">نسبة التحصيل</div>
                    <div class="kpi-value">{data['collection_rate']:.1f}%</div>
                </div>
            </div>
            
            <div class="card">
                <div class="section-title">تحليل الأداء</div>
                <table>
                    <tr>
                        <td style="font-weight: 600;">نسبة الإنجاز</td>
                        <td style="font-weight: 700;">{data['completion']:.0f}%</td>
                    </tr>
                    <tr>
                        <td style="font-weight: 600;">نسبة التحصيل</td>
                        <td style="font-weight: 700;">{data['collection_rate']:.1f}%</td>
                    </tr>
                    <tr>
                        <td style="font-weight: 600;">الوحدات المباعة</td>
                        <td style="font-weight: 700;">{data['sold_units']} من {data['total_units']}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: 600;">المتبقي من المبيعات</td>
                        <td style="font-weight: 700;">{format_number(data['remaining_sales'])} ر.س</td>
                    </tr>
                </table>
            </div>
        </div>
        
        <footer>
            <strong>حي السعادة 132</strong> · شركة قصر القصور · الرياض
            <br>
            <small>تم التحديث: {data['timestamp']}</small>
        </footer>
    </div>
    
    <script>
        setInterval(function() {{ location.reload(); }}, 10000);
    </script>
</body>
</html>"""
    
    return html

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
